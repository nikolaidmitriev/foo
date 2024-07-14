from os import walk
from openpyxl import load_workbook, Workbook

# default configuration__________________________________________________________________
# all files have to be in the same directory
# openpyxl have to be installed in python
selected_a_points = [0.01, 0.05, 0.1, 0.5, 1, 3, 5]
colums_for_points = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H'] # this has to be more by one

# functions______________________________________________________________________________
def get_points(some_sheet):
    variable_list = []
    j = 3
    while (some_sheet['B' + str(j)].value or some_sheet['C' + str(j)].value) != None:
        u = some_sheet['B' + str(j)].value
        i = some_sheet['C' + str(j)].value
        variable_list.append([u, i, True])
        j += 1
    return variable_list

def filter_processing(variable_list):
    for point in range(0, len(variable_list) - 2):
        if variable_list[point + 1][2] == False:
            continue
        coord1 = [variable_list[point][0], variable_list[point][1]]
        coord2 = [variable_list[point + 1][0], variable_list[point + 1][1]]
        coord3 = [variable_list[point + 2][0], variable_list[point + 2][1]]
        length12 = (coord2[0] - coord1[0])**2 + (coord2[1] - coord1[1])**2
        length13 = (coord3[0] - coord1[0])**2 + (coord3[1] - coord1[1])**2
        length23 = (coord3[0] - coord2[0])**2 + (coord3[1] - coord2[1])**2
        condition = (length12 > length13) or (length23 > length13)
        if condition:
            variable_list[point + 2][2] = False
    return variable_list

def search_a(variable_list, point_search):
    for point in range(0, len(variable_list)):
        if point_search > variable_list[point][1]:
            continue
        else:
            prev = variable_list[point - 1]
            after = variable_list[point]
            if prev[2] == False:
                prev = variable_list[point - 2]
            if after[2] == False:
                if point == (len(variable_list) - 1):
                    return [variable_list[-3], variable_list[-2]]
                after = variable_list[point + 1]
            return [prev, after]
    return [variable_list[-2], variable_list[-1]]

def linear_calc(selected_i_points, variable_list):
    calc_v = []
    for sel_i in selected_i_points:
        line = search_a(variable_list, sel_i)
        point = (line[1][0] - line[0][0])*(sel_i - line[0][1])/(line[1][1] - line[0][1]) + line[0][0]
        calc_v.append(round(point, 3))
    return(calc_v)

def create_result_sheet(lst):
    result = []
    for i in range(0, len(dir_lst_filt)):
        wb = load_workbook(filename = dir_lst_filt[i])
        sheet = wb[str(wb.sheetnames[0])]
        v_a_points = filter_processing(get_points(sheet))
        v_points = linear_calc(selected_a_points, v_a_points)
        v_points.append(dir_lst_filt[i][:-5])
        result.append(v_points)
        print(dir_lst_filt[i] + ' ... done!')
    return result

def writing_sheet(arr, col, sheet):
    for i in range(0, len(arr)):
        for j in range(0, len(col)):
            row = col[j] + str(i + 1)
            sheet[row] = arr[i][j]

# code process___________________________________________________________________________
resulted_workbook = Workbook()
resulted_worksheet = resulted_workbook.active

dir_lst = next(walk('.'))[2]
dir_lst.sort()
dir_lst_filt = list(filter(lambda k: 'xlsx' in k, dir_lst))

# if script used before, ignore previous results
if ('results.xlsx' in dir_lst_filt):
    dir_lst_filt.remove('results.xlsx')

resulted_internal_arr = create_result_sheet(dir_lst_filt)

writing_sheet(resulted_internal_arr, colums_for_points, resulted_worksheet)

resulted_workbook.save('results.xlsx')

print('all conversions done successfully')
print(str(len(dir_lst_filt)),' files processed')
print('press ENTER to exit')

input()
